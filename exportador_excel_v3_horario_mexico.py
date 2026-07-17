"""Exportador Excel del Sistema de Tiendas Escolares CECyTEA V3.

Este módulo NO recalcula pagos ni adeudos. Recibe la salida de
``procesador_v3.py`` y crea un Excel administrativo a partir de esa fuente de
verdad.

Hojas principales:
- Reporte Ejecutivo
- Adeudos
- Detalle de Cobranza
- Trazabilidad

Cuando existen movimientos con claves fuera del catálogo activo, se agrega una
hoja adicional: ``Movimientos por Revisar``. Así ningún depósito se pierde en
silencio.
"""

from __future__ import annotations

from datetime import datetime
from zoneinfo import ZoneInfo
from pathlib import Path
from typing import Mapping

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportes_v3 import generar_reportes

AZUL = "2B247C"
VERDE = "C6EFCE"
AMARILLO = "FFEB9C"
ROJO = "FFC7CE"
GRIS = "D9E1F2"
BLANCO = "FFFFFF"
BORDE = Side(style="thin", color="D9E1F2")

NOMBRES_HOJAS_BASE = [
    "Reporte Ejecutivo",
    "Adeudos",
    "Detalle de Cobranza",
    "Trazabilidad",
]


def _normalizar_fechas_para_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Convierte fechas de pandas a valores compatibles con Excel."""
    salida = df.copy()
    for columna in salida.columns:
        if "FECHA" in str(columna).upper() or str(columna).upper() == "MES":
            if pd.api.types.is_datetime64_any_dtype(salida[columna]):
                salida[columna] = salida[columna].dt.tz_localize(None)
    return salida


def _aplicar_estilo_general(ws) -> None:
    """Da formato uniforme a encabezados, filtros, bordes y anchos."""
    encabezados = list(ws.iter_rows(min_row=1, max_row=1))
    if encabezados:
        for cell in encabezados[0]:
            cell.fill = PatternFill("solid", fgColor=AZUL)
            cell.font = Font(color=BLANCO, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=BORDE, right=BORDE, top=BORDE, bottom=BORDE)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=BORDE, right=BORDE, top=BORDE, bottom=BORDE)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 28

    for columna in range(1, ws.max_column + 1):
        letra = get_column_letter(columna)
        maximo = 0
        for fila in range(1, min(ws.max_row, 250) + 1):
            valor = ws.cell(row=fila, column=columna).value
            if valor is not None:
                maximo = max(maximo, len(str(valor)))
        ws.column_dimensions[letra].width = min(max(maximo + 3, 12), 38)


def _formato_monetario(ws, encabezados: list[str]) -> None:
    """Aplica moneda a las columnas que contienen importes."""
    for indice, encabezado in enumerate(encabezados, start=1):
        texto = str(encabezado).upper()
        if any(token in texto for token in ("ADEUDO", "SALDO", "ESPERADO", "PAGADO", "PENDIENTE", "MONTO", "TOTAL")):
            for fila in range(2, ws.max_row + 1):
                ws.cell(row=fila, column=indice).number_format = '$#,##0.00'


def _formato_fechas(ws, encabezados: list[str]) -> None:
    for indice, encabezado in enumerate(encabezados, start=1):
        texto = str(encabezado).upper()
        if texto == "MES":
            for fila in range(2, ws.max_row + 1):
                ws.cell(row=fila, column=indice).number_format = 'mmm-yy'
        elif "FECHA" in texto:
            for fila in range(2, ws.max_row + 1):
                ws.cell(row=fila, column=indice).number_format = 'dd/mm/yyyy'


def _pintar_adeudos(ws, encabezados: list[str]) -> None:
    """Colorea el estado general y los importes de adeudo sin alterar datos."""
    mapa = {str(valor).upper(): indice + 1 for indice, valor in enumerate(encabezados)}
    if "ESTADO_GENERAL" in mapa:
        col = get_column_letter(mapa["ESTADO_GENERAL"])
        rango = f"{col}2:{col}{ws.max_row}"
        ws.conditional_formatting.add(
            rango,
            CellIsRule(operator="equal", formula=['"CON ADEUDO"'], fill=PatternFill("solid", fgColor=ROJO)),
        )
        ws.conditional_formatting.add(
            rango,
            CellIsRule(operator="equal", formula=['"SALDO A FAVOR"'], fill=PatternFill("solid", fgColor=VERDE)),
        )
        ws.conditional_formatting.add(
            rango,
            CellIsRule(operator="equal", formula=['"AL CORRIENTE"'], fill=PatternFill("solid", fgColor=GRIS)),
        )


def _pintar_ejecutivo(ws) -> None:
    """Da una lectura visual al estado mensual del reporte ejecutivo."""
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            texto = str(cell.value or "").upper()
            if "PENDIENTE" in texto:
                cell.fill = PatternFill("solid", fgColor=ROJO)
            elif "PARCIAL" in texto:
                cell.fill = PatternFill("solid", fgColor=AMARILLO)
            elif "PAGADO" in texto:
                cell.fill = PatternFill("solid", fgColor=VERDE)
            elif "SIN CUOTA" in texto:
                cell.fill = PatternFill("solid", fgColor=GRIS)

    for fila in range(2, ws.max_row + 1):
        ws.row_dimensions[fila].height = 42
    ws.freeze_panes = "D2"
    for letra in ("A", "B", "C"):
        ws.column_dimensions[letra].width = 20 if letra != "B" else 34


def _crear_resumen(resultados: Mapping[str, pd.DataFrame], adeudos: pd.DataFrame) -> pd.DataFrame:
    """Crea una hoja breve de control sin recalcular los importes operativos."""
    resumen = resultados["resumen_planteles"]
    fecha = datetime.now(ZoneInfo("America/Mexico_City")).strftime("%d/%m/%Y %H:%M")
    return pd.DataFrame(
        [
            ["Fecha de generación", fecha],
            ["Planteles activos procesados", int(len(resumen))],
            ["Planteles con adeudo", int((adeudos["ADEUDO_TOTAL"] > 0.005).sum())],
            ["Planteles con saldo a favor", int((adeudos["SALDO_FAVOR_TOTAL"] > 0.005).sum())],
            ["Adeudo total", float(adeudos["ADEUDO_TOTAL"].sum())],
            ["Saldo a favor total", float(adeudos["SALDO_FAVOR_TOTAL"].sum())],
        ],
        columns=["INDICADOR", "VALOR"],
    )


def exportar_excel_v3(
    resultado_procesamiento: Mapping[str, pd.DataFrame],
    destino: str | Path,
) -> Path:
    """Crea el Excel final a partir de la salida de ``procesar_periodo``.

    El archivo generado no vuelve a aplicar FIFO ni modifica saldos: solo exporta
    y presenta los resultados ya calculados.
    """
    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)

    reportes = generar_reportes(resultado_procesamiento)
    adeudos = reportes["adeudos"]
    resumen = _crear_resumen(resultado_procesamiento, adeudos)

    hojas: list[tuple[str, pd.DataFrame]] = [
        ("Resumen", resumen),
        ("Reporte Ejecutivo", reportes["reporte_ejecutivo"]),
        ("Adeudos", adeudos),
        ("Detalle de Cobranza", reportes["detalle_cobranza"]),
        ("Trazabilidad", resultado_procesamiento["trazabilidad"]),
    ]

    no_reconocidos = resultado_procesamiento.get("movimientos_no_reconocidos")
    if isinstance(no_reconocidos, pd.DataFrame) and not no_reconocidos.empty:
        hojas.append(("Movimientos por Revisar", no_reconocidos))

    with pd.ExcelWriter(destino, engine="openpyxl") as writer:
        for nombre_hoja, df in hojas:
            _normalizar_fechas_para_excel(df).to_excel(writer, sheet_name=nombre_hoja, index=False)

    libro = load_workbook(destino)
    for nombre_hoja in libro.sheetnames:
        ws = libro[nombre_hoja]
        encabezados = [cell.value for cell in ws[1]]
        _aplicar_estilo_general(ws)
        _formato_monetario(ws, encabezados)
        _formato_fechas(ws, encabezados)

        if nombre_hoja == "Adeudos":
            _pintar_adeudos(ws, encabezados)
        elif nombre_hoja == "Reporte Ejecutivo":
            _pintar_ejecutivo(ws)
        elif nombre_hoja == "Resumen":
            ws.column_dimensions["A"].width = 32
            ws.column_dimensions["B"].width = 24
            for fila in range(2, ws.max_row + 1):
                indicador = ws.cell(row=fila, column=1).value
                if indicador in {"Adeudo total", "Saldo a favor total"}:
                    ws.cell(row=fila, column=2).number_format = '$#,##0.00'

    libro.save(destino)
    return destino
