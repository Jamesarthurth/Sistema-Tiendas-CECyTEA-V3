"""Pruebas del Paso 6: creación del Excel final V3."""

from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory

import pandas as pd
from openpyxl import load_workbook

from exportador_excel_v3 import exportar_excel_v3
from lector_global import normalizar_global
from motor_v3 import MachoteV3
from procesador_v3 import procesar_periodo


def crear_machote_prueba() -> MachoteV3:
    planteles = pd.DataFrame([
        {"CLAVE_PLANTEL": "CJM", "NOMBRE_PLANTEL": "CECyTEA Jesús María", "ESTATUS_TIENDA": "ACTIVA"},
        {"CLAVE_PLANTEL": "AST", "NOMBRE_PLANTEL": "CECyTEA Asientos", "ESTATUS_TIENDA": "ACTIVA"},
    ])
    tarifas = pd.DataFrame([
        {"CLAVE_PLANTEL": "CJM", "NOMBRE_PLANTEL": "CECyTEA Jesús María", "MES": "2026-02-01", "CUOTA": 1000.0, "EE": 100.0},
        {"CLAVE_PLANTEL": "CJM", "NOMBRE_PLANTEL": "CECyTEA Jesús María", "MES": "2026-03-01", "CUOTA": 1000.0, "EE": 100.0},
        {"CLAVE_PLANTEL": "AST", "NOMBRE_PLANTEL": "CECyTEA Asientos", "MES": "2026-02-01", "CUOTA": 500.0, "EE": 50.0},
    ])
    tarifas["MES"] = pd.to_datetime(tarifas["MES"])
    saldos = pd.DataFrame([
        {"CLAVE_PLANTEL": "CJM", "NOMBRE_PLANTEL": "CECyTEA Jesús María", "SALDO_FAVOR_CUOTA": 0.0, "SALDO_FAVOR_EE": 0.0},
        {"CLAVE_PLANTEL": "AST", "NOMBRE_PLANTEL": "CECyTEA Asientos", "SALDO_FAVOR_CUOTA": 0.0, "SALDO_FAVOR_EE": 0.0},
    ])
    return MachoteV3(planteles=planteles, tarifas=tarifas, saldos_iniciales=saldos)


def crear_global_prueba() -> pd.DataFrame:
    return pd.DataFrame([
        {
            "  Fecha": "2026-02-10",
            "NÚMERO": "REC-001",
            "Matrícula": "CJM9900001",
            "Nombre(s)": "CECYTEA TIENDA ESCOLAR",
            "CUOTA RECUPERACION": 1500.0,
            "OTROS INGRESOS (ENERGIA ELEC)": 150.0,
        },
        {
            "  Fecha": "2026-02-11",
            "NÚMERO": "REC-002",
            "Matrícula": "XYZ9900001",
            "Nombre(s)": "TIENDA ESCOLAR SIN CATALOGO",
            "CUOTA RECUPERACION": 999.0,
            "OTROS INGRESOS (ENERGIA ELEC)": 9.0,
        },
    ])


def main() -> None:
    movimientos = normalizar_global(crear_global_prueba())
    procesado = procesar_periodo(crear_machote_prueba(), movimientos)

    with TemporaryDirectory() as temporal:
        salida = exportar_excel_v3(procesado, Path(temporal) / "Reporte_V3_Prueba.xlsx")
        assert salida.exists()

        libro = load_workbook(salida, data_only=False)
        esperadas = {
            "Resumen",
            "Reporte Ejecutivo",
            "Adeudos",
            "Detalle de Cobranza",
            "Trazabilidad",
            "Movimientos por Revisar",
        }
        assert esperadas.issubset(set(libro.sheetnames))

        detalle = libro["Detalle de Cobranza"]
        encabezados_detalle = [detalle.cell(row=1, column=i).value for i in range(1, detalle.max_column + 1)]
        assert encabezados_detalle == [
            "CLAVE_PLANTEL", "NOMBRE_PLANTEL", "MES", "CONCEPTO", "ESPERADO",
            "PAGADO", "PENDIENTE", "ESTADO", "FECHA_ULTIMO_ABONO",
        ]
        assert detalle.max_row == 7  # 6 renglones de detalle + encabezado
        assert detalle.freeze_panes == "A2"

        adeudos = libro["Adeudos"]
        encabezados_adeudos = [adeudos.cell(row=1, column=i).value for i in range(1, adeudos.max_column + 1)]
        assert "ADEUDO_TOTAL" in encabezados_adeudos
        assert "DETALLE_ADEUDO" not in encabezados_adeudos

        revisar = libro["Movimientos por Revisar"]
        assert revisar.max_row == 2  # encabezado + clave XYZ

    print("OK · Prueba 6: se genera un Excel final con las hojas administrativas requeridas.")
    print("OK · Prueba 6: Detalle de Cobranza conserva su estructura como fuente de verdad.")
    print("OK · Prueba 6: movimientos fuera del catálogo quedan visibles para revisión.")
    print("\nPRUEBAS DE EXPORTACIÓN EXCEL SUPERADAS.\n")


if __name__ == "__main__":
    main()
